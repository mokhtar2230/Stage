import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import mysql.connector
from mysql.connector import Error
import hashlib
def connect_to_database():
    """Connect to the MySQL database and return the connection."""
    try:
        conn = mysql.connector.connect(
            host='localhost',        
            user='root',            
            password='admin123',       
            database='primatec'    
        )
        if conn.is_connected():
            print("Successfully connected to the database")
            return conn
        else:
            print("Failed to connect to the database")
            return None
    except Error as e:
        print(f"Error: {e}")
        return None

def check_filename_in_db(filename):
    """Check if the filename exists in the database."""
    conn = connect_to_database()
    if conn:
        try:
            cursor = conn.cursor()
            query = 'SELECT nom FROM `n/a` WHERE nom = %s'
            cursor.execute(query, (filename,))
            result = cursor.fetchone()
        except Error as e:
            print(f"Database error: {e}")
            result = None
        finally:
            cursor.close()
            conn.close()
        return result is not None
    return False

def hash_content(content):
    """Calculate the MD5 hash of the given content."""
    hasher = hashlib.md5()
    hasher.update(content)
    return hasher.hexdigest()

def check_failed_files(input_file):
    """Check if the file exists in the Failed table and compare file content."""
    
    # Extract the filename from the input file path
    filename = os.path.basename(input_file)

    conn = connect_to_database()
    if conn:
        try:
            cursor = conn.cursor()
            
            # Query to check if the filename exists and retrieve its content
            query = 'SELECT file_data FROM Failed WHERE file_name = %s'
            cursor.execute(query, (filename,))
            result = cursor.fetchone()

            if result:
                # Fetch the content from the database (as a byte object)
                db_file_content = result[0]
                
                # Calculate the hash of the provided file's content
                with open(input_file, 'rb') as file:
                    file_content = file.read()

                # Debug: Print sizes and hash values
                print(f"Database file content size: {len(db_file_content)} bytes")
                print(f"Input file content size: {len(file_content)} bytes")
                
                db_file_hash = hash_content(db_file_content)
                file_hash = hash_content(file_content)
                print(f"Database file hash: {db_file_hash}")
                print(f"Input file hash: {file_hash}")

                # Compare the hashes of the database content and the input file content
                if db_file_hash == file_hash:
                    print("File content matches the one in the database.")
                    return True
                else:
                    print("File content does not match the one in the database.")
                    return False
            else:
                print("Filename not found in the Failed table.")
                return False
        except Error as e:
            print(f"Database error: {e}")
            return False
        finally:
            cursor.close()
            conn.close()
    else:
        print("Failed to connect to the database.")
        return False
    

    
def find_column_index(sheet, column_name):
    """Find the index of the column by name in the sheet."""
    for row in sheet.iter_rows(min_row=1):
        for cell in row:
            if cell.value == column_name:
                return cell.column, cell.row
    raise ValueError(f"Column '{column_name}' not found")

def set_column_widths(sheet):
    """Set column widths based on the length of the longest value in each column."""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

def process_file(input_file, output_sheet, result_col_index):
    """Process the input file and write the results to the output sheet."""
    success_fill = PatternFill(start_color='c6efce', end_color='92bf8d', fill_type='solid')  
    failed_fill = PatternFill(start_color='ffa9a3', end_color='92bf8d', fill_type='solid')  
    No_Results_fill = PatternFill(start_color='ffeb9c', end_color='FF0000', fill_type='solid')    
    na_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')       
    result_font = Font(color='000000')  # Black text

    file_name = os.path.splitext(os.path.basename(input_file))[0]
    error_message = ""

    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        try:
            routing_status_col_index, routing_status_row_index = find_column_index(sheet, "Routing status")
            correct_payload_routed_col_index, correct_payload_routed_row_index = find_column_index(sheet, "Correct payload Routed ")
            latency_col_index, latency_row_index = find_column_index(sheet, "Latency (ms)")
            SW_routed_col_index, SW_routed_row_index = find_column_index(sheet, "SW") 
            SW_Value_routed_col_index = SW_routed_col_index + 1
            SW_Value_routed_value = sheet.cell(row=SW_routed_row_index, column=SW_Value_routed_col_index).value

            analysis_comment_col_index, _ = find_column_index(output_sheet, "Analysis Comment")
            keyword_analysis_col_index, _ = find_column_index(output_sheet, "TC Keyword Analysis (SWBug;NewSWBug; SW Restriction; HW Restriction; In analysis; Test Infrastructure Restriction; Feature Not Implemented; TC Error)")
            Domain_col_index, _ = find_column_index(output_sheet, "Domain")
            SW_version_col_index, _ = find_column_index(output_sheet, "SW version")
            ECU_Tested_col_index, _ = find_column_index(output_sheet, "ECU_Tested")
            TC_ID_col_index, _ = find_column_index(output_sheet, "TC ID")
            TC_Name_col_index, _ = find_column_index(output_sheet, "TC_Name")
        except ValueError as e:
            error_message = f"Error finding columns in '{file_name}': {e}"
            print(error_message)
            return "No Results", error_message

        file_name_parts = file_name.split("_")
        ECU_Tested = file_name_parts[0]
        TC_Name = file_name_parts[1] + "_" + file_name_parts[2]
        TC_ID = TC_Name + "_" + ECU_Tested
        Domain = "Routing"

        if check_filename_in_db(TC_ID):
            result = "N/A"
            additional_info = "No routes in K-Matrix for this test case"
            keyword_analysis = "SW Restriction"
        elif check_failed_files(input_file):
            result = "failed"
            additional_info = ""
            keyword_analysis = ""
           
        else:
            all_true_routing = True
            all_latency_ok = True

            for row in range(routing_status_row_index + 1, sheet.max_row + 1):
                routing_status_value = sheet.cell(row=row, column=routing_status_col_index).value
                correct_payload_routed_value = sheet.cell(row=row, column=correct_payload_routed_col_index).value
                latency_value = sheet.cell(row=row, column=latency_col_index).value

                if not (routing_status_value is None or routing_status_value == True) or not (correct_payload_routed_value is None or correct_payload_routed_value == True):
                    all_true_routing = False

                try:
                    latency_value = int(latency_value) if latency_value is not None else 0
                except ValueError:
                    latency_value = 0

                if latency_value > 3:
                    all_latency_ok = False

                if not all_true_routing or not all_latency_ok:
                    result = "No Results"
                    error_message = f"File '{file_name}', Row {row}: Routing Status={routing_status_value}, Correct Payload Routed={correct_payload_routed_value}, Latency={latency_value}"
                    break
            else:
                result = "Success" if all_true_routing and all_latency_ok else "No Results"

            additional_info = ""
            keyword_analysis = ""

        last_row = output_sheet.max_row + 1

        output_sheet.cell(row=last_row, column=ECU_Tested_col_index, value=ECU_Tested)
        output_sheet.cell(row=last_row, column=TC_ID_col_index, value=TC_ID)
        output_sheet.cell(row=last_row, column=TC_Name_col_index, value=TC_Name)

        result_cell = output_sheet.cell(row=last_row, column=result_col_index, value=result)
        if result == "Success":
            result_cell.fill = success_fill
        elif result == "No Results":
            result_cell.fill = No_Results_fill
        elif result == "failed":
            result_cell.fill = failed_fill
        else:
            result_cell.fill = na_fill
        result_cell.font = result_font

        output_sheet.cell(row=last_row, column=analysis_comment_col_index, value=additional_info)
        output_sheet.cell(row=last_row, column=keyword_analysis_col_index, value=keyword_analysis)
        output_sheet.cell(row=last_row, column=Domain_col_index, value=Domain)
        output_sheet.cell(row=last_row, column=SW_version_col_index, value=SW_Value_routed_value)
        print(check_failed_files(input_file))
        if result == "No Results":
            print(f"Error message: {error_message}")

        print(f"Processed '{file_name}' with result: {result}")
        return result, error_message

    except (FileNotFoundError, PermissionError, IOError) as e:
        error_message = f"Error processing file '{file_name}': {e}"
        print(error_message)
        return "No Results", error_message

def taux(sheet):
    """Calculate the number of successful, No Results, and 'N/A' results in the sheet."""
    success_count = 0
    No_Results_count = 0
    na_count = 0
    failed_count = 0

    result_col_index = find_column_index(sheet, "Result (Passed, Failed, No Results, No Execution, N/A)")[0]

    for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
        result_value = sheet.cell(row=row, column=result_col_index).value
        if result_value == "Success":
            success_count += 1
        elif result_value == "N/A":
            na_count += 1
        elif result_value == "failed":
            failed_count += 1
        elif result_value == "No Results":
            No_Results_count += 1

    return success_count, No_Results_count, na_count , failed_count
