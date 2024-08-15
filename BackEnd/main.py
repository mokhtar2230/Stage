import os
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory,send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font
from file_utils import create_output_directory, get_input_files, clear_upload_folder,insert_report
from excel_utils import process_file, set_column_widths, find_column_index, taux
from failed import getallfailed, add, remove,save_file
from na import getallna, addna, removena
from auth import login
from collections import defaultdict
from flask import Flask, request, jsonify, session
from flask_session import Session
from History import get_reports_by_user,get_reports
import logging
import base64
import io
from UserManager import afficheuser ,add_user,get_roles,removeuser
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Configuration for Flask-Session
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'your_prefix:'
app.config['SESSION_FILE_DIR'] = './flask_session/'  

Session(app)
CORS(app, supports_credentials=True)


@app.route('/login', methods=['POST'])
def login_route():
    print(session)  
    
    return login()

@app.route  ('/afficheuser', methods=['GET'])
def users():
    try:
        result = afficheuser()
        if result:
            return jsonify({'message': 'Data found in the database', 'data': result}), 200
        else:
            return jsonify({'message': 'No data found in the database'}), 404
    except Exception as e:
        print(f"Error retrieving data: {e}")
        return jsonify({'error': 'Error retrieving data'}), 500



@app.route('/add_user', methods=['POST'])
def adduser():

        return add_user()

@app.route('/roles', methods=['GET'])
def getroles():
    return get_roles()


@app.route('/session', methods=['GET'])
def get_session():
    user_id = session.get('user_id')
    if user_id:
        return jsonify({"user_id": user_id}), 200
    else:
        return jsonify({"message": "No active session"}), 404


@app.route('/upload/', methods=['POST'])
def upload_file():
    input_folder = 'uploads'
    clear_upload_folder(input_folder)
    
    if 'files' not in request.files:
        return jsonify({'error': 'No files part'}), 400
    
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No selected files'}), 400
    
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)

    for file in files:
        if file.filename == '':
            continue  

        file_path = os.path.join(input_folder, file.filename)
        file.save(file_path)
    
    output_dir = create_output_directory()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'result_{timestamp}.xlsx')

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    header_row = [
        "Domain", "Requirement_ID", "ECU_Tested", "TC ID", "TC_Name", 
        "TC_Description", "Result (Passed, Failed, No Results, No Execution, N/A)", 
        "Analysis Comment", "Defect ID", "KPM Ticket ID", 
        "TC Keyword Analysis (SWBug;NewSWBug; SW Restriction; HW Restriction; In analysis; Test Infrastructure Restriction; Feature Not Implemented; TC Error)",
        "Defect Description", "SW Bug Severity", "SW version", "Traces"
    ]
    header_fill = PatternFill(start_color='3366CC', end_color='3366CC', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_num, value in enumerate(header_row, start=2):
        cell = output_sheet.cell(row=1, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font

    output_sheet.title = 'Raw_Data'

    result_col_index = find_column_index(output_sheet, "Result (Passed, Failed, No Results, No Execution, N/A)")[0]
    data = []
    errors = []
    ECU_set = set()
    TC_name_set = set()

    for input_file in get_input_files(input_folder):
        result, error_message = process_file(input_file, output_sheet, result_col_index)
        if error_message:
            errors.append(error_message)

        if result != "N/A":
            filename = os.path.splitext(os.path.basename(input_file))[0]
            parts = filename.split("_")
            if len(parts) >= 3:
                ECU = parts[0]
                TC_name = f"{parts[1]}_{parts[2]}"
            elif len(parts) == 2:
                ECU = parts[0]
                TC_name = parts[1]
            else:
                ECU = parts[0]
                TC_name = "Unknown_TC"

            ECU_set.add(ECU)
            TC_name_set.add(TC_name)
            data.append(f"{filename}§{result}")

    set_column_widths(output_sheet)
    output_workbook.save(output_file)
    
    user_id = session.get('user_id')
    print(f"The user ID is: {user_id}")
    insert_report(user_id, output_file)

    ECU_result_counts = defaultdict(lambda: {"No Results": 0, "Success": 0, "Failed": 0})
    TC_name_result_counts = defaultdict(lambda: {"No Results": 0, "Success": 0, "Failed": 0})

    for item in data:
        filename, result = item.split("§")
        parts = filename.split("_")
        ECU_Tested = parts[0]
        TC_name = f"{parts[1]}_{parts[2]}" if len(parts) >= 3 else "Unknown_TC"

        if result == "No Results":
            ECU_result_counts[ECU_Tested]["No Results"] += 1
            TC_name_result_counts[TC_name]["No Results"] += 1
        elif result == "Success":
            ECU_result_counts[ECU_Tested]["Success"] += 1
            TC_name_result_counts[TC_name]["Success"] += 1
        elif result == "failed":
            ECU_result_counts[ECU_Tested]["Failed"] += 1
            TC_name_result_counts[TC_name]["Failed"] += 1

    ECU_percentages = {
        ECU_Tested: {
            "No Results": (counts["No Results"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100,
            "Success": (counts["Success"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100,
            "Failed": (counts["Failed"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100
        }
        for ECU_Tested, counts in ECU_result_counts.items()
    }

    TC_name_percentages = {
        TC_name: {
            "No Results": (counts["No Results"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100,
            "Success": (counts["Success"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100,
            "Failed": (counts["Failed"] / (counts["No Results"] + counts["Success"] + counts["Failed"])) * 100
        }
        for TC_name, counts in TC_name_result_counts.items()
    }

    try:
        wb = openpyxl.load_workbook(output_file)
        sheet = wb['Raw_Data']
        success_count, No_Results_count, na_count, failed_count = taux(sheet)
    except Exception as e:
        print(f"Error processing results: {e}")
        return jsonify({'error': 'Error processing results'}), 500

    total_count = success_count + No_Results_count + na_count + failed_count
    pr_percentage = (success_count / total_count * 100) if total_count > 0 else 0
    pr_no_Results = (No_Results_count / total_count * 100) if total_count > 0 else 0
    pr_na = (na_count / total_count * 100) if total_count > 0 else 0
    pr_failed = (failed_count / total_count * 100) if total_count > 0 else 0
    print(os.path.basename(output_file))
  
    response = {
        'message': f"Files processed successfully. Output file path: {output_file}",
        'success_count': success_count,
        'No_Results_count': No_Results_count,
        'na_count': na_count,
        'failed_count': failed_count,    
        'pr_failed': pr_failed,
        'pr': pr_percentage,
        'pr_no_Results': pr_no_Results,
        'pr_na': pr_na,
        'output_file_url': f"http://127.0.0.1:5000//download/{os.path.basename(output_file)}",
        'error_message': errors,
        'ECU': [
            {
                'name': ECU_Tested,
                'No_Results': round(percents['No Results'], 1),
                'success': round(percents['Success'], 1),
                'failed': round(percents['Failed'], 1)
            }
            for ECU_Tested, percents in ECU_percentages.items()
        ],
        'TC_Name': [
            {
                'name': TC_name,
                'No_Results': round(percents['No Results'], 1),
                'success': round(percents['Success'], 1),
                'failed': round(percents['Failed'], 1)
            }
            for TC_name, percents in TC_name_percentages.items()
        ]          
    }

    if errors:
        response['errors'] = errors

    return jsonify(response), 200


@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    output_dir = create_output_directory()  
    return send_from_directory(directory=output_dir, path=filename, as_attachment=True)




@app.route('/getallfailed/', methods=['GET'])
def getFailed_all():
    

   
    failed = getallfailed()
    
    if failed is None:
        return jsonify({'error': 'Error fetching reports'}), 500

    # Format and return the reports
    formatted_failed = [
        {
           
            'file_name': failed['file_name'],
            'id': failed['id'],
            'file_data': failed['file_data'],
        }
        for failed in failed
    ]
    
    return jsonify({'faileds': formatted_failed}), 200





@app.route('/uploadFailed/', methods=['POST'])
def upload_failed():
    upload_folder = 'uploads'

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    try:
        # Save the file and get its path
        file_path = save_file(file, upload_folder)

        # Insert file data into the database
        add(file_path)

        return jsonify({'message': 'File uploaded successfully'}), 200
    except Exception as e:
        print(f"Error uploading file: {e}")
        return jsonify({'error': 'Error uploading file'}), 500






@app.route('/removeuser/', methods=['DELETE'])
def remove_user():
    data = request.get_json()
    if not data or 'id' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    try:
        removeuser(data)
        return jsonify({'message': 'Record removed successfully'}), 200
    except Exception as e:
        print(f"Error removing record: {e}")
        return jsonify({'error': 'Error removing record'}), 500  





@app.route('/remove/', methods=['DELETE'])
def remove_record():
    data = request.get_json()
    if not data or 'id' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    try:
        remove(data)
        return jsonify({'message': 'Record removed successfully'}), 200
    except Exception as e:
        print(f"Error removing record: {e}")
        return jsonify({'error': 'Error removing record'}), 500

@app.route('/getallna/', methods=['GET'])
def get_allna():
    try:
        result = getallna()
        if result:
            return jsonify({'message': 'Data found in the database', 'data': result}), 200
        else:
            return jsonify({'message': 'No data found in the database'}), 404
    except Exception as e:
        print(f"Error retrieving data: {e}")
        return jsonify({'error': 'Error retrieving data'}), 500

@app.route('/addna/', methods=['POST'])
def add_recordna():
    data = request.get_json()
    if not data or 'FileName' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    try:
        addna(data)
        return jsonify({'message': 'Record added successfully'}), 200
    except Exception as e:
        print(f"Error adding record: {e}")
        return jsonify({'error': 'Error adding record'}), 500

@app.route('/removena/', methods=['DELETE'])
def remove_recordna():
    data = request.get_json()
    if not data or 'id' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    try:
        removena(data)
        return jsonify({'message': 'Record removed successfully'}), 200
    except Exception as e:
        print(f"Error removing record: {e}")
        return jsonify({'error': 'Error removing record'}), 500


@app.route('/logout', methods=['POST'])
def logout():
    session.pop('user_id', None)  # Remove user_id from session
    return jsonify({"message": "Logout successful"}), 200
@app.route('/check_session/', methods=['GET'])
def check_session():
    user_id = session.get('user_id')
    if user_id:
        return jsonify({'message': 'Session is active', 'user_id': user_id}), 200
    else:
        return jsonify({'message': 'No active session'}), 404
    


@app.route('/reports/', methods=['GET'])
def get_user_reports():
    user_id = session.get('user_id')

    if not user_id:
        return jsonify({'error': 'User not logged in'}), 401

    # Fetch reports from the database
    reports = get_reports_by_user(user_id)
    
    if reports is None:
        return jsonify({'error': 'Error fetching reports'}), 500

    # Format and return the reports
    formatted_reports = [
        {
            'id': report['id'],
            'username': report['username'],
            'file_name': report['file_name'],
            'file_data': report['file_data'],
            'user_id': report['user_id'],   
        }
        for report in reports
    ]
    
    return jsonify({'reports': formatted_reports}), 200







@app.route('/get_reports/', methods=['GET'])
def get__reports():
    user_id = session.get('user_id')

    if not user_id:
        return jsonify({'error': 'User not logged in'}), 401

    # Fetch reports from the database
    reports = get_reports(user_id)
    
    if reports is None:
        return jsonify({'error': 'Error fetching reports'}), 500

    # Format and return the reports
    formatted_reports = [
        {
            'id': report['id'],
            'username': report['username'],
            'file_name': report['file_name'],
            'file_data': report['file_data'],
            'user_id': report['user_id'],   
        }
        for report in reports
    ]
    
    return jsonify({'reports': formatted_reports}), 200





@app.route('/download_report/<int:user_id>/<int:report_id>', methods=['GET'])
def download_report(user_id, report_id):
    reports = get_reports_by_user(user_id)
    if not reports:
        return jsonify({"error": "No reports found for the user"}), 404

    report = next((r for r in reports if r['id'] == report_id), None)
    if not report:
        return jsonify({"error": "Report not found"}), 404

    if 'file_data' not in report or not report['file_data']:
        return jsonify({"error": "No file data available for the report"}), 404

    # Decode the Base64 encoded file data
    file_data = base64.b64decode(report['file_data'])
    file_name = report['file_name']

    # Create an in-memory bytes buffer
    buffer = io.BytesIO(file_data)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=file_name
    )

    

if __name__ == "__main__":
    app.run(debug=True)
